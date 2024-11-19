import pandas as pd
import requests
from PIL import Image
from io import BytesIO
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import tempfile

def download_and_resize_image(url, max_height_inches=2):
    try:
        # Download image from URL
        response = requests.get(url)
        img = Image.open(BytesIO(response.content))
        
        # Check if image is portrait (height > width) and rotate if needed
        if img.height > img.width:
            # Rotate image 90 degrees clockwise
            img = img.transpose(Image.Transpose.ROTATE_90)
        
        # Convert inches to pixels (assuming 96 DPI)
        max_height_pixels = int(max_height_inches * 96)
        
        # Calculate new dimensions maintaining aspect ratio
        aspect_ratio = img.width / img.height
        new_height = min(max_height_pixels, img.height)
        new_width = int(new_height * aspect_ratio)
        
        # Resize image
        img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
        temp_path = temp_file.name
        img.save(temp_path)
        
        return temp_path
    except Exception as e:
        print(f"Error processing image from {url}: {str(e)}")
        return None

def process_excel_file(file_path):
    # Get absolute file paths
    abs_file_path = os.path.abspath(file_path)
    file_name, file_ext = os.path.splitext(abs_file_path)
    new_file_path = f"{file_name}-updated{file_ext}"
    
    # Copy original file to new file
    wb = load_workbook(abs_file_path)
    sheet = wb.worksheets[1]  # Get second sheet
    
    # Get maximum row count
    max_row = sheet.max_row
    
    # Add header for image column
    image_col = get_column_letter(sheet.max_column + 1)
    sheet[f"{image_col}1"] = "image"
    
    # Process each row except header
    for row in range(2, max_row + 1):
        # Get URL from first column
        url = sheet[f"A{row}"].value
        
        if url:
            print(f"Processing row {row}: {url}")  # Debug print
            # Download and resize image

             #if temp_image_path is not a url
            if not url.startswith("https://"):
                print(f"Error processing image from {url}")
                break
            
           

            temp_image_path = download_and_resize_image(url)
            
            if temp_image_path:
                try:
                    # Add image to Excel
                    img = XLImage(temp_image_path)
                    cell = f"{image_col}{row}"
                    sheet.add_image(img, cell)
                    
                    # Adjust row height to fit image
                    sheet.row_dimensions[row].height = 96  # approximately 1 inch
                except Exception as e:
                    print(f"Error adding image to Excel for row {row}: {str(e)}")
                #finally:
                    # Clean up temporary file
                    #try:
                    #    os.unlink(temp_image_path)
                    #except Exception as e:
                    #    print(f"Error deleting temporary file: {str(e)}")
    
    # Adjust column width
    sheet.column_dimensions[image_col].width = 30  # Adjust as needed
    
    # Save the workbook
    try:
        wb.save(new_file_path)
        print(f"File saved as: {new_file_path}")
    except Exception as e:
        print(f"Error saving workbook: {str(e)}")

# Usage
if __name__ == "__main__":
    # Get the current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(current_dir, "Data-QnA-Academic.xlsx")  # Replace with your file name
    process_excel_file(excel_file)